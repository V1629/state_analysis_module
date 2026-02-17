"""
Chat Logger with Excel Export
Now supports state activation status - shows N/A for inactive states
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os


class ChatLogger:
    """
    Logs chat messages with emotional state to Excel file
    
    Structure:
    - Message
    - Impact Score
    - Current Chat (Short-term, Mid-term, Long-term) - Top 1 emotion each
    - User Profile (Short-term, Mid-term, Long-term) - Top 1 emotion each
    - State Activation Status
    
    NEW: Shows "N/A" for states that aren't activated yet
    """

    def __init__(self, file_path="chat_logs.xlsx"):
        self.file_path = file_path
        self.ensure_workbook()

    def ensure_workbook(self):
        """Create Excel file if it doesn't exist"""
        if not os.path.exists(self.file_path):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Chat Logs"
            
            # Set up headers
            column_headers = [
                "Timestamp",
                "Message",
                "Impact Score",
                # Current Chat - Top Emotion (1 each)
                "Current_ST_Emotion",
                "Current_ST_Score",
                "Current_MT_Emotion",
                "Current_MT_Score",
                "Current_LT_Emotion",
                "Current_LT_Score",
                # User Profile - Top Emotion (1 each)
                "Profile_ST_Emotion",
                "Profile_ST_Score",
                "Profile_MT_Emotion",
                "Profile_MT_Score",
                "Profile_LT_Emotion",
                "Profile_LT_Score",
                # State Activation Status
                "MT_Status",
                "LT_Status",
                "Profile_Age_Days",
                "Message_Count",
            ]
            
            worksheet.append(column_headers)
            
            # Style headers
            header_fill_color = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_style = Font(bold=True, color="FFFFFF")
            
            for header_cell in worksheet[1]:
                header_cell.fill = header_fill_color
                header_cell.font = header_font_style
                header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Set column widths
            worksheet.column_dimensions["A"].width = 25  # Timestamp
            worksheet.column_dimensions["B"].width = 45  # Message
            worksheet.column_dimensions["C"].width = 15  # Impact Score
            
            for column_index in range(4, 18):
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = 18
            
            # Set row height for header
            worksheet.row_dimensions[1].height = 30
            
            workbook.save(self.file_path)

    def log_chat(
        self,
        message: str,
        impact_score: float,
        current_state: dict,
        profile_state: dict,
        activation_status: dict = None,
        profile_age_days: int = 0,
        message_count: int = 0
    ):
        """
        Log a chat message with current and profile states
        
        Args:
            message: User's message
            impact_score: Emotional impact score
            current_state: Dict with 'short_term', 'mid_term', 'long_term' 
                          each containing list of (emotion, score) tuples
            profile_state: Dict with 'short_term', 'mid_term', 'long_term'
                          each containing list of (emotion, score) tuples
            activation_status: Dict showing which states are active
            profile_age_days: Age of profile in days
            message_count: Total messages processed
        """
        workbook = openpyxl.load_workbook(self.file_path)
        worksheet = workbook.active
        
        # Prepare row data
        row_data_values = [
            datetime.now().isoformat(),
            message,
            round(impact_score, 4),
        ]
        
        # Add current chat top emotion (1 for each state)
        for temporal_state_type in ['short_term', 'mid_term', 'long_term']:
            top_emotion_data = self._extract_top_emotion(current_state.get(temporal_state_type, []))
            row_data_values.extend(top_emotion_data)
        
        # Add profile top emotion (1 for each state)
        for temporal_state_type in ['short_term', 'mid_term', 'long_term']:
            top_emotion_data = self._extract_top_emotion(profile_state.get(temporal_state_type, []))
            row_data_values.extend(top_emotion_data)
        
        # Add state activation status
        mid_term_activation_status = self._get_activation_status('mid_term', activation_status)
        long_term_activation_status = self._get_activation_status('long_term', activation_status)
        
        row_data_values.extend([
            mid_term_activation_status,
            long_term_activation_status,
            profile_age_days,
            message_count
        ])
        
        # Append row
        worksheet.append(row_data_values)
        
        # Style the new row
        current_row_number = worksheet.max_row
        for column_index in range(1, len(row_data_values) + 1):
            current_cell = worksheet.cell(row=current_row_number, column=column_index)
            current_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            current_cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color code N/A cells
            if current_cell.value == "N/A":
                current_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                current_cell.font = Font(color="9C0006")
            elif current_cell.value in ["✅ ACTIVE", "⏳ INACTIVE"]:
                if current_cell.value == "✅ ACTIVE":
                    current_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    current_cell.font = Font(color="006100")
                else:
                    current_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    current_cell.font = Font(color="9C6500")
        
        # Set row height
        worksheet.row_dimensions[current_row_number].height = 25
        
        workbook.save(self.file_path)

    @staticmethod
    def _extract_top_emotion(emotion_list: list) -> list:
        """
        Extract top 1 emotion and score from emotion list
        Returns ["N/A", "N/A"] if list is empty or contains N/A
        
        Args:
            emotion_list: List of tuples [(emotion, score), ...]
        
        Returns:
            List [emotion, score] or ["N/A", "N/A"]
        """
        if emotion_list and len(emotion_list) > 0:
            emotion, score = emotion_list[0]
            
            # Check if it's already N/A placeholder
            if emotion == "N/A":
                return ["N/A", "N/A"]
            
            return [emotion, round(float(score), 4)]
        else:
            return ["N/A", "N/A"]

    @staticmethod
    def _get_activation_status(state_type: str, activation_status: dict = None) -> str:
        """
        Get activation status for a state
        
        Args:
            state_type: 'mid_term' or 'long_term'
            activation_status: Dict from get_state_activation_info()
        
        Returns:
            "✅ ACTIVE" or "⏳ INACTIVE" or "N/A"
        """
        if activation_status is None:
            return "N/A"
        
        if state_type not in activation_status:
            return "N/A"
        
        is_active = activation_status[state_type].get('is_active', False)
        
        return "✅ ACTIVE" if is_active else "⏳ INACTIVE"